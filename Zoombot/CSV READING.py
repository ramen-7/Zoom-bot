import openpyxl as xl
from datetime import datetime


min = int(datetime.now().strftime("%M"))
r_min = round(min/10)*10
now = datetime.now().strftime(f"%H:{r_min}:00")
col = 0
ro = 0
day = datetime.now().strftime("%A").upper()
print(now)
wb = xl.load_workbook("PythonTimetable.xlsx")
sheet = wb['PythonTimetable']
for column in range(2, sheet.max_column + 1):
    cell1 = sheet.cell(1, column)
    if day == cell1.value:
        col = column
        print(col)

for row in range(2, sheet.max_row + 1):
    cell2 = sheet.cell(row, 1)
    if now == cell2.value:
        ro = row
        print(ro)