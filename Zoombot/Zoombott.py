import subprocess
import pyautogui
import time
import openpyxl as xl
from datetime import datetime
import os

def sign_in(meeting_id):
    zoomprocess = subprocess.Popen(["C:\\Users\\shvmt\\AppData\\Roaming\\Zoom\\bin\\Zoom.exe"]) #location of zoom on your pc
    if zoomprocess.returncode == 0:
        print("Zoom opened successfully")
    else:
        print("Error in opening Zoom")
    time.sleep(7)
    join_btn = pyautogui.locateCenterOnScreen("JoinButton.png")
    print(join_btn)
    pyautogui.moveTo(join_btn)
    pyautogui.click()
    time.sleep(7)
    print("Join pressed")

    pyautogui.write(meeting_id)
    print("Writing...")

    #dnc_audio_btn = pyautogui.locateCenterOnScreen("DoNotConnectToAudio.png")
    #pyautogui.moveTo(dnc_audio_btn)
    #pyautogui.click()
    #print("Success3")
    #time.sleep(1)

    dnc_video_btn = pyautogui.locateCenterOnScreen("TurnOffMyVideo.png")
    pyautogui.moveTo(dnc_video_btn)
    pyautogui.click()
    print("Video Turned Off")
    time.sleep(1)

    join_f_btn = pyautogui.locateCenterOnScreen("Join.png")
    pyautogui.moveTo(join_f_btn)
    pyautogui.click()
    print("Class joined")
    time.sleep(15)

    connect_audio_btn = pyautogui.locateCenterOnScreen("ConnectMyAuido.png")
    pyautogui.moveTo(connect_audio_btn)
    pyautogui.click()
    print("Audio Connected")

def leave_meet():
    dnc_leave_meeting_btn = pyautogui.locateCenterOnScreen("LeaveButton.png")
    pyautogui.moveTo(dnc_leave_meeting_btn)
    pyautogui.click()
    print("Left meeting")
    time.sleep(1)

def time_check():
    meeting_connection_id = 0
    mini = int(datetime.now().strftime("%M"))
    r_min = round(mini/10)*10
    if r_min == 0:
        r_min = str("00")
    now = datetime.now().strftime(f"%H:{r_min}")
    col = 0
    ro = 0
    day = datetime.now().strftime("%A").upper()
    print(now)
    wb = xl.load_workbook("PythonTimetable.xlsx")
    sheet = wb['PythonTimetable']

    for column in range(2, sheet.max_column + 1):
        cell1 = sheet.cell(1, column)
        if day == cell1.value:
            col = column
            print(f"col = {col}")

    for row in range(2, sheet.max_row + 1):
        cell2 = sheet.cell(row, 1)
        if now == cell2.value:
            ro = row
            print(f"row = {ro}")
            break

    if ro == 0:
        print('There is no class at this time')
        exit(1)

    #if meeting_connection_id !=
    meet_id = sheet.cell(ro, col)
    print(f"meet_id =  {meet_id.value}")
    sign_in(str(meet_id.value))



if __name__ == "__main__":
    time_check()
